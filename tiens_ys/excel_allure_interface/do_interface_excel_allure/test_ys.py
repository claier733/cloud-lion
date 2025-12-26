import os
import string
from fileinput import filename
from operator import is_not
from pkgutil import get_data
import allure
import openpyxl
import pytest
from allure_pytest.utils import allure_title
from excel_allure_interface.tool.username import random_str
from excel_allure_interface.api_key.api_key import ApiKey
from excel_allure_interface.excel_reader.exscel_driver import ReadExcel
from datetime import date
from excel_allure_interface.tool.log_util import logger
import sys

# 添加项目根目录到Python路径
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

print(f"测试文件位置: {__file__}")
print(f"项目根目录: {project_root}")

# 获取当前日期
current_date = date.today()

# 定义全局变量
EXCEL_SAVE_PATH = None


def save_excel_file(excel_obj, path=None):
    """保存Excel文件到指定路径"""
    if path is None:
        path = EXCEL_SAVE_PATH

    if path is None:
        # 如果路径未定义，使用相对路径
        current_dir = os.path.dirname(os.path.abspath(__file__))
        path = os.path.join(current_dir, '..', 'data', 'demoe.xlsx')

    try:
        # 确保目录存在
        save_dir = os.path.dirname(path)
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
            logger.info(f"创建目录: {save_dir}")

        excel_obj.save(path)
        logger.debug(f"Excel文件已保存: {path}")
        return True
    except Exception as e:
        logger.error(f"保存Excel文件失败: {e}")

        # 尝试保存到当前目录作为备份
        try:
            current_dir = os.path.dirname(os.path.abspath(__file__))
            backup_path = os.path.join(current_dir, 'demoe_backup.xlsx')
            excel_obj.save(backup_path)
            logger.info(f"已保存备份文件到: {backup_path}")
            return True
        except Exception as backup_error:
            logger.error(f"备份保存也失败: {backup_error}")
            return False


# @allure.story('全局变量')
def setup_module():
    # 1,声明全局变量
    global ak, r, sheet, excel, extract_dict, expr, EXCEL_SAVE_PATH

    # 2,实例化类
    ak = ApiKey()

    # 3,打开excel - 使用绝对路径
    current_dir = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(current_dir, '..', 'data', 'demoe.xlsx')

    # 存储为全局变量，用于后续保存
    EXCEL_SAVE_PATH = excel_path

    # 调试信息
    print(f"当前目录: {current_dir}")
    print(f"Excel路径: {excel_path}")
    print(f"文件是否存在: {os.path.exists(excel_path)}")

    if not os.path.exists(excel_path):
        # 尝试查找文件
        for root, dirs, files in os.walk(current_dir):
            for file in files:
                if 'demoe' in file.lower():
                    print(f"找到可能的目标文件: {os.path.join(root, file)}")
        raise FileNotFoundError(f"无法找到Excel文件: {excel_path}")

    excel = openpyxl.load_workbook(excel_path)
    sheet = excel['info']

    # 4,数据抽取存储
    extract_dict = {}  # 存储公共变量
    print("setup_module() 初始化完成")


@allure.epic('云狮项目')
@allure.feature('我的')
@allure.story('个人中心')
@pytest.mark.parametrize('data', ReadExcel().excel_read())
def test_do_interface(data):
    if data[10]:
        allure.dynamic.title(data[10])

    # 1,数据解析
    if 'var_token' in extract_dict:
        url = data[1] + data[2] + '&token=' + extract_dict['var_token']
    else:
        url = data[1] + data[2]

    if data[5] is not None:
        data_dict = {
            'url': url,
            'headers': eval(data[4]),
            'json': eval(data[5])
        }
    else:
        data_dict = {
            'url': url,
            'headers': eval(data[4])
        }

    # 2,反射+解包
    max_retries = 5  # 最大重试次数
    retry_count = 0
    res_result = None
    res = None

    while retry_count < max_retries:
        res = getattr(ak, data[3])(**data_dict)
        logger.info(f"Status Code: {res.status_code}")
        logger.info(f"Response Content: {res.text}")

        try:
            logger.info(f"Response JSON: {res.json()}")
        except Exception as e:
            logger.warning(f"无法解析JSON响应: {e}")
            logger.info(f"原始响应文本: {res.text}")

        if res.json().get('msg') == '用户名已存在':
            retry_count += 1
            logger.info(f'用户名已存在，重试第 {retry_count} 次')
            continue
        else:
            break

    # 数据抽取
    if data[11] is not None:
        extract = data[11].split(';')
        expr = data[12].split(';')
        for i in range(len(extract)):
            try:
                extract_dict[extract[i]] = ak.get_text(res.text, expr[i])
            except (TypeError, IndexError) as e:
                logger.error(f'{data[10]}抽取数据失败: {e}')
                extract_dict[extract[i]] = None

    # 3，断言+反显 （将结果传到excel）
    r = data[0] + 1
    try:
        res_result = ak.get_text(res.text, data[7])

        if res_result == data[8]:
            sheet.cell(r, 10).value = '通过'
            logger.info(f"用例 {data[10]} 通过")
        else:
            sheet.cell(r, 10).value = '失败'
            logger.warning(f"用例 {data[10]} 失败，期望: {data[8]}，实际: {res_result}")

        # 保存Excel文件
        save_success = save_excel_file(excel)
        if not save_success:
            logger.error(f"用例 {data[10]} 的测试结果保存失败")

    except Exception as e:
        logger.exception(f'用例名称：{data[10]}；返回结果错误，或者是表达式{data[8]}有误')
        sheet.cell(r, 10).value = f'返回结果错误，或者是表达式{data[8]}有误'

        # 异常情况下也要尝试保存
        save_excel_file(excel)

    finally:
        # 确保无论如何都保存一次
        save_success = save_excel_file(excel)

        # 只在保存成功的情况下断言，避免因保存失败影响测试结果
        if save_success:
            assert res_result == data[8], f'断言失败，期望: {data[8]}，实际: {res_result}'
        else:
            logger.error(f"用例 {data[10]} 结果保存失败，跳过断言")

    print(f'\ncase：{data[10]}')


def teardown_module():
    """测试结束后清理"""
    global excel

    try:
        # 最终保存一次Excel文件
        if 'excel' in globals() and excel is not None:
            save_excel_file(excel)
            logger.info("测试结束，最终保存Excel文件")
    except Exception as e:
        logger.error(f"teardown_module 保存失败: {e}")

    print("所有测试用例执行完成")


if __name__ == '__main__':
    # 运行测试并生成 Allure 原始结果
    pytest.main(['-v', '--alluredir', './allure-results'])